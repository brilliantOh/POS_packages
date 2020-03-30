# receipt_draft.py

import pandas as pd
import openpyxl as xl
from initialize_reform import POS_calculator, cal
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows


# receive order and save data(excel)
class Receipt:
    def __init__(self):
        self.details_cols_list = ['주문일시', '메뉴', '수량', '단가', '금액', '지불수단']
        self.receipt_df = pd.DataFrame(columns=self.details_cols_list)

        self.summary_cols_list = ['주문일시', '금액']
        self.summary_df = pd.DataFrame(columns=self.summary_cols_list)

        self.wb_filename = 'order_history.xlsx'

        try:
            self.wb = xl.load_workbook(filename=self.wb_filename)
            self.details_ws = self.wb['receipt']
            self.summary_ws = self.wb['summary']
        except FileNotFoundError:
            self.wb = xl.Workbook(self.wb_filename)
            self.details_ws = self.wb.active
            self.details_ws.title = 'receipt'
            self.details_ws.append(self.details_cols_list)
            self.summary_ws = self.wb.create_sheet()
            self.summary_ws.title = 'summary'
            self.summary_ws.append(self.summary_cols_list)
            self.wb.save(filename=self.wb_filename)

    def receive_order(self, payment_method_str):
        self.receipt_df = pd.DataFrame(columns=self.details_cols_list)

        datetime_now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for i in range(len(cal.cart_list)):
            if cal.cart_list[i] not in cal.cart_list[:i]:
                details_list = [datetime_now_str, cal.cart_list[i], cal.return_menu_qty(cal.cart_list[i]),
                                cal.return_menu_cost(cal.cart_list[i]), cal.return_menu_amount(cal.cart_list[i]),
                                payment_method_str]
                self.receipt_df = self.receipt_df.append(pd.DataFrame([details_list], columns=self.details_cols_list),
                                                         ignore_index=True)

        cal.cancel_all_qty()
        self.create_summary_df()
        self.save_receipt_df_to_excel()
        self.save_summary_df_to_excel()

        return self.receipt_df

    def create_summary_df(self):
        self.summary_df = pd.DataFrame(columns=self.summary_cols_list)

        summary_list = [self.receipt_df['주문일시'][0], sum(self.receipt_df['금액'])]
        self.summary_df = self.summary_df.append(pd.DataFrame([summary_list], columns=self.summary_cols_list),
                                                 ignore_index=True)
        return self.summary_df

    def save_receipt_df_to_excel(self):
        for row in dataframe_to_rows(self.receipt_df, index=False, header=False):
            self.details_ws.append(row)
            self.wb.save(filename=self.wb_filename)

    def save_summary_df_to_excel(self):
        for row in dataframe_to_rows(self.summary_df, index=False, header=False):
            self.summary_ws.append(row)
            self.wb.save(filename=self.wb_filename)


rec = Receipt()
