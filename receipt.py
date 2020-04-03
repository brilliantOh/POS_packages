'''
receipt.py
======================
order management module of POS_packages
'''

import pandas as pd
import openpyxl as xl
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from calculator import POS_calculator, cal


# receive order and save data(excel)
class POS_receipt:
    '''POS의 주문 기능, 주문내역 저장(excel) 기능을 하는 클래스입니다.

    '''
    def __init__(self):
        '''생성자
        :explain: order_history.xlsx 파일이 경로에 이미 있다면 불러오고, 없다면 새로 생성합니다.
        receipt/summary worksheet를 불러오거나, 생성해서 지정합니다.
        '''
        self.details_cols_list = ['주문일시', '메뉴', '수량', '단가', '금액', '지불수단']
        self.receipt_df = pd.DataFrame(columns=self.details_cols_list)

        self.summary_cols_list = ['주문일시', '금액']
        self.summary_df = pd.DataFrame(columns=self.summary_cols_list)

        self.wb_filename = r'C:\Users\JH\POS_packages\order_history.xlsx'

        try:
            self.wb = xl.load_workbook(filename=self.wb_filename)
            self.details_ws = self.wb['receipt']
            self.summary_ws = self.wb['summary']
        except FileNotFoundError:
            self.wb = xl.Workbook(self.wb_filename)
            self.details_ws = self.wb.active
            self.details_ws.title = 'receipt'
            self.details_ws.append(self.details_cols_list)
            self.summary_ws = self.wb.create_sheet()
            self.summary_ws.title = 'summary'
            self.summary_ws.append(self.summary_cols_list)
            self.wb.save(filename=self.wb_filename)

    def receive_order(self, payment_method_str):
        '''
        :explain: 입력한 지불수단에 따른 주문내역 dataframe을 생성하고 반환합니다. 장바구니 비우기, 요약내역 생성, excel에 저장하는 함수를 호출합니다.

        :param string payment_method_str: 지불수단
        :return dataframe: receipt_df 주문내역
        '''
        self.receipt_df = pd.DataFrame(columns=self.details_cols_list)

        datetime_now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for i in range(len(cal.cart_list)):
            if cal.cart_list[i] not in cal.cart_list[:i]:
                details_list = [datetime_now_str, cal.cart_list[i], cal.return_menu_qty(cal.cart_list[i]),
                                cal.return_menu_cost(cal.cart_list[i]), cal.return_menu_amount(cal.cart_list[i]),
                                payment_method_str]
                self.receipt_df = self.receipt_df.append(pd.DataFrame([details_list], columns=self.details_cols_list),
                                                         ignore_index=True)

        cal.cancel_all_qty()
        self.create_summary_df()
        self.save_receipt_df_to_excel()
        self.save_summary_df_to_excel()

        return self.receipt_df

    def create_summary_df(self):
        '''
        :explain: 요약내역 dataframe을 생성하고 반환합니다.

        :return dataframe: summary_df 요약내역
        '''
        self.summary_df = pd.DataFrame(columns=self.summary_cols_list)

        summary_list = [self.receipt_df['주문일시'][0], sum(self.receipt_df['금액'])]
        self.summary_df = self.summary_df.append(pd.DataFrame([summary_list], columns=self.summary_cols_list),
                                                 ignore_index=True)
        return self.summary_df

    def save_receipt_df_to_excel(self):
        '''
        :explain: 주문내역 dataframe을 excel로 저장합니다.

        '''
        for row in dataframe_to_rows(self.receipt_df, index=False, header=False):
            self.details_ws.append(row)
            self.wb.save(filename=self.wb_filename)

    def save_summary_df_to_excel(self):
        '''
        :explain: 요약내역 dataframe을 excel로 저장합니다.

        '''
        for row in dataframe_to_rows(self.summary_df, index=False, header=False):
            self.summary_ws.append(row)
            self.wb.save(filename=self.wb_filename)


rec = POS_receipt()